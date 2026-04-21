"""Convert the n8n official nodes xlsx into `data/nodes/catalog_discovery.json`.

Reads `n8n_official_nodes_reference.xlsx` at the project root, walks the three
main sheets (Core Nodes, AI LangChain Nodes, App Action Nodes) and writes a
JSON array of `NodeCatalogEntry`-shaped records to
`data/nodes/catalog_discovery.json`.

Usage (run from repo root or the scripts/ dir):

    python scripts/xlsx_to_catalog.py

The xlsx header columns are:
    節點名稱 (Display Name) | type 欄位值 | 類別 | 說明

App Trigger Nodes and the JSON 結構參考 sheet are intentionally skipped for the
MVP; the former lacks a `類別`/`說明` column, the latter is documentation.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# Make backend/ importable when invoked from repo root.
_THIS = Path(__file__).resolve()
_PROJECT_ROOT = _THIS.parents[1]
sys.path.insert(0, str(_PROJECT_ROOT / "backend"))

from app.models import NodeCatalogEntry  # noqa: E402

XLSX_PATH = _PROJECT_ROOT / "n8n_official_nodes_reference.xlsx"
OUT_PATH = _PROJECT_ROOT / "data" / "nodes" / "catalog_discovery.json"

SHEETS = ["Core Nodes", "AI LangChain Nodes", "App Action Nodes"]

_STOPWORDS_EN = {
    "the", "a", "an", "and", "or", "to", "of", "in", "on", "for",
    "with", "by", "is", "are", "be", "as", "at", "from", "it", "this",
    "that", "into", "via", "using", "use", "can", "will", "nodes", "node",
}
_TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9]+|[\u4e00-\u9fff]+")


def _keywords(text: str, display_name: str, max_k: int = 5) -> list[str]:
    """Very small tokenizer: English alnum + CJK runs; dedupe; cap at 5."""
    if not text:
        text = ""
    pool = f"{display_name} {text}"
    seen: list[str] = []
    for tok in _TOKEN_RE.findall(pool):
        low = tok.lower()
        if low in _STOPWORDS_EN:
            continue
        if tok in seen:
            continue
        seen.append(tok)
        if len(seen) >= max_k:
            break
    return seen


def _normalise_entry(row: tuple, category_fallback: str) -> NodeCatalogEntry | None:
    display_name = (row[0] or "").strip()
    node_type = (row[1] or "").strip()
    category = (row[2] or category_fallback or "Uncategorised").strip()
    description = (row[3] or "").strip() or display_name
    if not node_type or not display_name:
        return None
    return NodeCatalogEntry(
        type=node_type,
        display_name=display_name,
        category=category,
        description=description,
        default_type_version=None,
    )


def load_entries(xlsx_path: Path = XLSX_PATH) -> list[NodeCatalogEntry]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    entries: dict[str, NodeCatalogEntry] = {}
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[warn] sheet not found: {sheet_name}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # header row is rows[0]
        for row in rows[1:]:
            # pad to 4 columns
            padded = tuple(list(row) + [None] * (4 - len(row)))
            entry = _normalise_entry(padded, category_fallback=sheet_name)
            if entry is None:
                continue
            if entry.type in entries:
                # prefer the one with a non-empty description
                if len(entry.description) > len(entries[entry.type].description):
                    entries[entry.type] = entry
                continue
            entries[entry.type] = entry
    return list(entries.values())


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[error] xlsx not found: {XLSX_PATH}", file=sys.stderr)
        return 2
    entries = load_entries()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = []
    for e in entries:
        d = e.model_dump()
        # enrich with derived keywords (not part of NodeCatalogEntry; kept at json level
        # for downstream RAG use).
        d["keywords"] = _keywords(e.description, e.display_name)
        payload.append(d)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[ok] wrote {len(payload)} entries to {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
